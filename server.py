#!/usr/bin/env python3
"""
PDF订单 → 生产单转换工具 服务器
- 静态文件服务（index.html、PDF.js、SheetJS 等）
- POST /export  接收 JSON 数据，返回带格式的 Excel 文件
"""
import http.server
import json
import os
import io
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── 模板格式常量（从生产单模板读取） ──────────────────────────
HEADER_BG      = "FF87CEFA"   # 天蓝色背景
HEADER_FONT    = "微软雅黑"
DATA_FONT      = "微软雅黑"
FONT_SIZE      = 9
ROW_HEIGHT     = 30

THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 表头填充 & 字体
HEADER_FILL = PatternFill(fill_type="solid", fgColor=HEADER_BG)
HEADER_FONT_NORMAL = Font(name=HEADER_FONT, size=FONT_SIZE, color="FFFFFFFF")  # 白色
DATA_FONT_NORMAL   = Font(name=DATA_FONT, size=FONT_SIZE)
DATA_FONT_WRAP     = Font(name=DATA_FONT, size=FONT_SIZE)

# 表头对齐（多数居中，少数左对齐）
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
# 数据行对齐（需要换行的列）
ALIGN_WRAP   = Alignment(vertical="center", wrap_text=True)

# ── 43列定义：[中文表头, 列宽, 表头对齐, 数据wrap] ────────────
COLUMNS = [
    # 列号  中文表头                      列宽      表头对齐       数据wrap
    ("序号",                              4.33,  ALIGN_CENTER,  False),
    ("店铺",                             17.66,  ALIGN_CENTER,  False),
    ("订单号",                            15.50,  ALIGN_CENTER,  False),
    ("编号",                             10.66,  ALIGN_CENTER,  False),
    ("产品名称",                          25.33,  ALIGN_CENTER,  False),
    ("供电方式",                           7.33,  ALIGN_CENTER,  False),
    ("数量",                             10.16,  ALIGN_LEFT,    False),
    ("面料型号1(上或前)",                   14.16,  ALIGN_CENTER,  False),
    ("面料颜色1",                         11.50,  ALIGN_CENTER,  False),
    ("面料型号2(下或后)",                   14.16,  ALIGN_CENTER,  False),
    ("面料颜色2",                          8.16,  ALIGN_CENTER,  False),
    ("遮光率",                             5.83,  ALIGN_CENTER,  False),
    ("内装/外装",                          8.00,  ALIGN_CENTER,  False),
    ("客户测量尺寸宽度INCH",               18.16,  ALIGN_CENTER,  False),
    ("客户测量尺寸高度INCH",               18.16,  ALIGN_CENTER,  False),
    ("客户测量尺寸斜边INCH",               18.16,  ALIGN_CENTER,  False),
    ("电机选型",                           9.33,  ALIGN_CENTER,  False),
    ("Hub名称",                            7.33,  ALIGN_CENTER,  False),
    ("Hub数量",                            7.33,  ALIGN_CENTER,  False),
    ("罩壳",                              9.83,  ALIGN_CENTER,  False),
    ("罩壳颜色",                           7.33,  ALIGN_CENTER,  False),
    ("底梁",                             16.83,  ALIGN_CENTER,  False),
    ("底梁颜色",                           7.33,  ALIGN_CENTER,  False),
    ("太阳能板名称",                       15.33,  ALIGN_CENTER,  False),
    ("太阳能板数量",                       10.33,  ALIGN_CENTER,  False),
    ("遥控器型号",                         14.16,  ALIGN_CENTER,  False),
    ("遥控器数量",                          8.83,  ALIGN_CENTER,  False),
    ("正转反转",                            7.33,  ALIGN_CENTER,  False),
    ("马达左装右装",                        10.33,  ALIGN_CENTER,  False),
    ("遮光条",                             5.83,  ALIGN_CENTER,  False),
    ("延长线",                             5.83,  ALIGN_CENTER,  False),
    ("延长线数量",                          8.83,  ALIGN_CENTER,  False),
    ("电源线",                             5.83,  ALIGN_CENTER,  False),
    ("电源适配器",                         27.66,  ALIGN_CENTER,  False),
    ("电源适配器数量",                      12.16,  ALIGN_CENTER,  False),
    ("房间编号",                           21.66,  ALIGN_CENTER,  False),
    ("包边",                             17.66,  ALIGN_CENTER,  False),
    ("客户信息",                          38.66,  ALIGN_CENTER,  True),
    ("备注",                             40.66,  ALIGN_CENTER,  True),
    ("付款时间",                           12.50,  ALIGN_CENTER,  False),
    ("客户邮箱电话",                       10.33,  ALIGN_CENTER,  False),
    ("系统标识ID",                          9.00,  ALIGN_CENTER,  False),
    ("是否加急",                            7.33,  ALIGN_CENTER,  False),
]

# 数据行中，key与列的对应（顺序与 COLUMNS 一致）
KEY_ORDER = [
    "seq_no","store","order_no","no","product","power","qty",
    "fabric1","color1","fabric2","color2","blackout","mount",
    "width","height","diagonal","motor","hub_name","hub_qty",
    "cover","cover_color","beam","beam_color",
    "solar_name","solar_qty","remote","remote_qty",
    "rotation","motor_side","shade_strip","ext_wire","ext_wire_qty",
    "power_wire","adapter","adapter_qty",
    "room_no","edging","customer","remark","pay_date","contact",
    "sys_id","urgent",
]

# 数值类型的字段（存为数字而非字符串）
NUMERIC_KEYS = {"qty","hub_qty","solar_qty","remote_qty","ext_wire_qty","adapter_qty",
                "width","height","diagonal","urgent"}
DATE_KEYS    = {"pay_date"}


def build_excel(rows_data: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "生产单"

    # ── 表头行 ────────────────────────────────────────────
    ws.row_dimensions[1].height = ROW_HEIGHT
    for col_idx, (header, col_w, h_align, _wrap) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill    = HEADER_FILL
        cell.font    = HEADER_FONT_NORMAL
        cell.alignment = h_align
        cell.border  = THIN_BORDER
        # 列宽
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = col_w

    # ── 数据行 ─────────────────────────────────────────────
    for row_idx, row in enumerate(rows_data, 2):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
        for col_idx, (header, _w, _ha, do_wrap) in enumerate(COLUMNS, 1):
            key = KEY_ORDER[col_idx - 1]
            # 序号列自动填行号
            if key == "seq_no":
                raw = row_idx - 1
            else:
                raw = row.get(key, "")

            # 类型转换
            value = None
            if raw not in (None, ""):
                if key in NUMERIC_KEYS:
                    try:    value = float(raw) if "." in str(raw) else int(raw)
                    except: value = raw
                elif key in DATE_KEYS:
                    try:    value = datetime.strptime(str(raw)[:10], "%Y-%m-%d")
                    except: value = raw
                else:
                    value = str(raw)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT_NORMAL

            # 对齐：部分列左对齐或需要换行
            if do_wrap:
                cell.alignment = ALIGN_WRAP
            elif header in ("房间编号", "面料型号1(上或前)", "付款时间"):
                cell.alignment = ALIGN_LEFT

            # 日期格式
            if key in DATE_KEYS and isinstance(value, datetime):
                cell.number_format = "YYYY/M/D"

    # ── 冻结首行 ──────────────────────────────────────────
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTTP 请求处理 ─────────────────────────────────────────
class Handler(http.server.SimpleHTTPRequestHandler):

    def do_POST(self):
        if self.path == "/export":
            length = int(self.headers.get("Content-Length", 0))
            body   = self.rfile.read(length)
            try:
                payload  = json.loads(body)
                rows     = payload.get("rows", [])
                filename = payload.get("filename", "生产单.xlsx")
                xlsx_bytes = build_excel(rows)

                self.send_response(200)
                self.send_header("Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition",
                    f'attachment; filename*=UTF-8\'\'{filename}')
                self.send_header("Content-Length", str(len(xlsx_bytes)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(xlsx_bytes)

            except Exception as e:
                self.send_error(500, str(e))
        else:
            self.send_error(404)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, GET, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def log_message(self, fmt, *args):
        # 过滤掉 GET 静态文件的日志，只显示 POST
        if args and "POST" in str(args[0]):
            print(f"  → {args[0]} {args[1]}")


if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8899
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    with http.server.ThreadingHTTPServer(("0.0.0.0", port), Handler) as httpd:
        print(f"服务已启动：http://localhost:{port}")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n服务已停止")
